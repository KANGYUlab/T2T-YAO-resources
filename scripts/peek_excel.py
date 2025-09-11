import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python peek_excel.py <excel_path>")
        sys.exit(1)

    excel_path = Path(sys.argv[1]).expanduser().resolve()
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    max_rows_to_show = 200
    max_cols_to_show = 12

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows_to_show, values_only=True), start=1):
        row_vals = ["" if v is None else str(v) for v in row[:max_cols_to_show]]
        print(f"{row_idx}\t" + " | ".join(row_vals))


if __name__ == "__main__":
    main()


